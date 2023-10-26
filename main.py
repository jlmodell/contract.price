import pandas as pd
import polars as pl
import os
from rich import print
from datetime import datetime
from pymongo import MongoClient
import warnings
import yaml

def get_config():
    config_path = r"config.yaml"
    if not os.path.exists(config_path):
        config_path = r"C:\temp\contracts_config.yaml"
    
    assert os.path.exists(config_path), f"Could not find config file at {config_path}"
    
    with open(config_path, "r") as f:
        config = yaml.load(f, Loader=yaml.FullLoader)

    return config

config = get_config()

MONGODB_URI=config["MONGODB_URI"]

warnings.filterwarnings('ignore')

save_path = r"C:\My Documents\Converted Contracts"
if not os.path.exists(save_path):
    os.makedirs(save_path)

base = r"C:\temp"
if not os.path.exists(base):
    os.makedirs(base)

contract = os.path.join(base, "specialpricing.csv")
assert os.path.exists(contract), """
**********************************************************************
    
    Step 1: run `GET.CONTRACT.INFO` in ROI
          
        * Input Contract Number
        * Hit "Enter" and Hit "F3" to continue        
        
    Step 2: Type "y" and hit "Enter" to continue below
          
**********************************************************************\n"""

sales_end_date = datetime.today().replace(day=1) - pd.DateOffset(days=1)
sales_end_date = pd.to_datetime(sales_end_date)    
sales_start_date_py = sales_end_date - pd.DateOffset(months=24) + pd.DateOffset(days=1)  
sales_file = os.path.join(base, "sales.for.period.csv")
assert os.path.exists(sales_file), f"""
**********************************************************************
        
    Step 0: run `GET.SALES.FOR.MDB` in ROI 
        
        > Perform step 0 *ONCE A MONTH* on/about 1st of the month <

        Calculating dates for you to use...

        * Input Sales Start Date -> {sales_start_date_py: %m%d%y}  <- and Hit "Enter"
        * Input Sales End Date -> {sales_end_date: %m%d%y}  <- and Hit "Enter"
        * Accept `ALL` for customers and hit "Enter"
        * Accept `ALL` for items and hit "Enter"
        * Hit "F3" to continue
        
        (This should take 2-3 minutes as it is a large query)
      
**********************************************************************\n"""

def print_instructions_to_terminal():
    """
    Prints instructions to terminal.
    """
    sales_end_date = datetime.today().replace(day=1) - pd.DateOffset(days=1)
    sales_end_date = pd.to_datetime(sales_end_date)    
    sales_start_date_py = sales_end_date - pd.DateOffset(months=24) + pd.DateOffset(days=1)
    
    instructions = f"""
**********************************************************************
        
    Step 0: run `GET.SALES.FOR.MDB` in ROI 
        
        > Perform step 0 *ONCE A MONTH* on/about 1st of the month <

        Calculating dates for you to use...

        * Input Sales Start Date -> {sales_start_date_py: %m%d%y}  <- and Hit "Enter"
        * Input Sales End Date -> {sales_end_date: %m%d%y}  <- and Hit "Enter"
        * Accept `ALL` for customers and hit "Enter"
        * Accept `ALL` for items and hit "Enter"
        * Hit "F3" to continue
        
        (This should take 2-3 minutes as it is a large query)

    Step 1: run `GET.CONTRACT.INFO` in ROI
          
        * Input Contract Number
        * Hit "Enter" and Hit "F3" to continue        
        
    Step 2: Type "y" and hit "Enter" to continue below
          
**********************************************************************
    {datetime.now(): %m/%d/%Y %H:%M:%S} --    
        Are you ready to continue? (y/n):\t"""

    ready = False
    while not ready:
        r = input(instructions)
        if r.lower().startswith("y"):
            ready = True
    
    return ready

# if day is between 1-5 print a reminder to run GET.SALES.FOR.MDB
if datetime.today().day <= 5:
    print("""
                        *-*-*-*-*-*-*-*-*-*-*-*-*
                        *                       *
                        *  Make Sure to Re-Run  *                              
                        *  `GET.SALES.FOR.MDB`  *
                        *                       *          
                        *-*-*-*-*-*-*-*-*-*-*-*-*""")

ready = False
while not ready:
    ready = print_instructions_to_terminal()

client = MongoClient(MONGODB_URI)
db = client.bussepricing
costs_db = db.costs
items_db = db.items

def get_cost(item: str) -> float:
    doc = costs_db.find_one({
        "alias": item
    })
    if doc:
        return float(doc["cost"])
    return 0.0

def get_description(item: str) -> str:
    doc = items_db.find_one({
        "item": item
    })
    if doc:
        return doc["name"]
    return ""

contract = pd.read_csv(contract, header=None, dtype=str, infer_datetime_format=True)
contract.columns = [
    "contract",
    "item",
    "contract name",
    "start date",
    "end date",
    "min quantity",
    "uom",
    "unit price",
    "py_sold",
    "cy_sold",
    "cust nbr",
    "cust name",
]

contract = contract[contract["item"].isna() == False]

contract["item"] = contract["item"].astype(str)
contract["cust nbr"] = contract["cust nbr"].astype(str)
contract["min quantity"] = contract["min quantity"].astype(int)
contract["unit price"] = contract["unit price"].astype(float)

_contract = contract.to_dict(orient="records")
contract = [c for c in _contract if c["item"] != ""]

curr_contract = f"{contract[0]['contract']} ({contract[0]['contract name']})"
curr_customers = list(set([c["cust name"] for c in contract]))
curr_contract_customers = ", ".join(curr_customers[:1]) + ", ..." if len(curr_customers) > 1 else ", ".join(curr_customers)
cust_codes = list(set([c.split("*")[0] for c in list(set([c["cust nbr"] for c in contract]))]))
curr_contract_start_date = contract[0]["start date"]
curr_contract_end_date = contract[0]["end date"]

items = {c["item"]: {"unit_price": c["unit price"], "min_qty": c["min quantity"], "uom": c["uom"]} for c in contract}

sales_end_date = datetime.today().replace(day=1) - pd.DateOffset(days=1)
sales_end_date = pd.to_datetime(sales_end_date)
sales_start_date = sales_end_date - pd.DateOffset(months=12)
sales_end_date_py = sales_end_date - pd.DateOffset(months=12)
sales_start_date_py = sales_end_date - pd.DateOffset(months=24) + pd.DateOffset(days=1)


sales = pl.read_csv(
    sales_file,
    infer_schema_length=20000,
)

sales = sales.with_columns(
    pl.col("ITEM_PART_NBR").cast(pl.Utf8),
    pl.col("CUST_NBR").cast(pl.Utf8),
    pl.col("INV_SO_DATE").str.to_date("%m-%d-%y").keep_name(),
)

sales = sales.to_pandas()

constructor = {
    "Contract #": curr_contract,
    "Customers": curr_contract_customers,
    "Start Date": curr_contract_start_date,
    "End Date": curr_contract_end_date,
    "Commission %": 0.04, # "B5"    
    "Items": []
}

for item, values in items.items():
    curr_sales = sales[
            (sales["ITEM_PART_NBR"] == item)
            & (sales["CUST_NBR"].isin(cust_codes))
            & (sales["INV_SO_DATE"] >= sales_start_date)
            & (sales["INV_SO_DATE"] <= sales_end_date)
        ].copy()
    prev_year_sales = sales[
            (sales["ITEM_PART_NBR"] == item)
            & (sales["CUST_NBR"].isin(cust_codes))
            & (sales["INV_SO_DATE"] >= sales_start_date_py)
            & (sales["INV_SO_DATE"] <= sales_end_date_py)
        ].copy()        
    
    constructor["Items"].append([
        item,
        get_description(item),
        values["min_qty"],
        values["uom"],
        values["unit_price"],
        get_cost(item),
        f"=({get_cost(item)} + (J2 * {values['unit_price']}) + (L2 * {values['unit_price']}))",
        f"=({values['unit_price']}-{get_cost(item)} - (J2 * {values['unit_price']}) - (L2 * {values['unit_price']})) / {values['unit_price']}",
        curr_sales["QUANTITY"].sum(),
        curr_sales["SALES"].sum(),
        prev_year_sales["QUANTITY"].sum(),
        prev_year_sales["SALES"].sum(),        
    ])

def build_workbook(constructor: dict[str, any]):
    import openpyxl
    from openpyxl.styles import Border, Side

    def set_border(ws, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)    
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet.append([
        "Contract #",
        constructor["Contract #"],
        "",
        "",
        "",
        "",
        "",
        "",
        "Contract Start:",
        pd.to_datetime(constructor["Start Date"]).strftime("%m/%d/%Y"),
        "Contract End:",
        pd.to_datetime(constructor["End Date"]).strftime("%m/%d/%Y"),        
    ]) # 1

    sheet.append([
        "Customers",
        constructor["Customers"],
        "",
        "",
        "",
        "",
        "Freight Terms:",
        "FOB HAUPPAUGE",     
        "Commission %:",
        0.04,
        "Giveback %:",
        0.03
    ]) # 2

    for col_num in range(1, 13, 2):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = openpyxl.styles.Font(bold=True)
        cell = sheet.cell(row=2, column=col_num)
        cell.font = openpyxl.styles.Font(bold=True)

    for col_num in range(3, 13):
        cell = sheet.cell(row=1, column=col_num)
        cell.alignment = openpyxl.styles.Alignment(horizontal="right")
        cell = sheet.cell(row=2, column=col_num)
        cell.alignment = openpyxl.styles.Alignment(horizontal="right")

    sheet.append([]) # 3

    sheet.append([
        "Item",
        "Description",
        "Min Qty",
        "UOM",
        "Unit Price",
        "Unit Cost",
        "Loaded Cost",
        "GP %",
        "Current Year CS",
        "Current Year $",
        "Previous Year CS",
        "Previous Year $",        
    ]) # 4

    for col_num in range(1, 13):
        cell = sheet.cell(row=4, column=col_num)
        cell.font = openpyxl.styles.Font(bold=True)

    for item_details in constructor["Items"]:        
        sheet.append(item_details)        

    for row_num in range(5, 5+len(constructor["Items"])):
        cell = sheet.cell(row=row_num, column=5)
        cell.number_format = "$#,##0.00"

        cell = sheet.cell(row=row_num, column=6)
        cell.number_format = "$#,##0.00"

        cell = sheet.cell(row=row_num, column=7)
        cell.number_format = "$#,##0.00"

        cell = sheet.cell(row=row_num, column=8)
        cell.number_format = "0.00%"

        cell = sheet.cell(row=row_num, column=9)
        cell.number_format = "#,##0"

        cell = sheet.cell(row=row_num, column=10)
        cell.number_format = "$#,##0.00"

        cell = sheet.cell(row=row_num, column=11)
        cell.number_format = "#,##0"

        cell = sheet.cell(row=row_num, column=12)
        cell.number_format = "$#,##0.00"    
    
    f1 = sheet.cell(row=1, column=6)
    f1.number_format = "mm/dd/yyyy"
    
    h1 = sheet.cell(row=1, column=8)
    h1.number_format = "mm/dd/yyyy"
    
    j2 = sheet.cell(row=2, column=10)
    j2.number_format = "0.00%"
    j2.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    l2 = sheet.cell(row=2, column=12)
    l2.number_format = "0.00%"
    l2.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for column in sheet.columns:
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(column[0].column)  # Get column letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    sheet.column_dimensions["D"].width = 8
    sheet.column_dimensions["E"].width = 18
    sheet.column_dimensions["F"].width = 18
    sheet.column_dimensions["G"].width = 18
    sheet.column_dimensions["H"].width = 18
    sheet.column_dimensions["I"].width = 18
    sheet.column_dimensions["J"].width = 18
    sheet.column_dimensions["K"].width = 18
    sheet.column_dimensions["L"].width = 18
    
    sheet.append([])
    sheet.append([])
    sheet.append([])

    sheet.append([
        "",
        f"* costs as of {datetime.now().strftime('%m/%d/%Y')}"
    ])

    max_rows = 4 + len(constructor["Items"]) + 5

    sheet.print_area = f"A1:L{max_rows}"

    sheet.print_title_rows = "1:4"

    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.5
    sheet.page_margins.bottom = 0.5

    sheet.page_setup.orientation = "landscape"

    sheet.page_setup.fitToPage = True
    sheet.page_setup.fitToHeight = False
    sheet.page_setup.fitToWidth = 1

    set_border(sheet, f"A1:L{max_rows}")

    contract_number = constructor["Contract #"].split(" ")[0]
    filename = os.path.join(save_path, f"{contract_number}.xlsx")
    workbook.save(filename)

    return filename

def open_workbook(filename: str):
    import subprocess
    try:
        subprocess.Popen([filename], shell=True)
    except Exception as e:        
        print(f"Could not open {filename}")
        print(e)

if __name__ == "__main__":
    # build workbook and save filename
    filename = build_workbook(constructor)

    # open workbook after building it
    open_workbook(filename)